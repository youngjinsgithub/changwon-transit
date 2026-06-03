"""대중교통 전환율 시나리오별 유류비·CO₂ 절감 효과 분석.

자가용 통행의 5/10/15/20% 가 시내버스로 전환됐을 때
창원시가 얻는 연간 유류비 절감액과 온실가스 감축량을 시뮬레이션.

산출:
  - analysis/yeongjin_chart/conversion_savings.xlsx
      · 시트 1 '가정·출처' : 입력 계수와 인용 출처
      · 시트 2 '시나리오 결과' : 4개 시나리오 × 7개 지표
      · 시트 3 '차트' : 막대그래프 (연간 유류비 절감 + CO₂ 절감)
  - analysis/yeongjin_chart/conversion_savings.png
      · 보고서 삽입용 이미지 (matplotlib)
"""
from __future__ import annotations

import io
import sys
from pathlib import Path

import matplotlib.pyplot as plt
import matplotlib.ticker as mtick
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

PROJECT_ROOT = Path(__file__).resolve().parents[2]
OUT_DIR = PROJECT_ROOT / "analysis" / "yeongjin_chart"
OUT_XLSX = OUT_DIR / "conversion_savings.xlsx"
OUT_PNG = OUT_DIR / "conversion_savings.png"


# ──────────────────────────────────────────────────────────────────────
# 1. 가정·계수 (전부 공공통계 인용)
# ──────────────────────────────────────────────────────────────────────
ASSUMPTIONS = [
    # (항목, 값, 단위, 출처)
    ("창원시 자동차 등록대수",       697_261,  "대",
     "KOSIS DT_1YL20731 (국토교통부 자동차등록현황보고, 2025)"),
    ("승용차 비중",                  0.80,    "비율",
     "국토교통부 자동차등록통계 — 전국 승용차 비중 약 80%"),
    ("자가용 승용차 (계산)",        557_809,  "대",
     "= 등록대수 × 승용차 비중"),
    ("1대당 일평균 주행거리",        33.0,    "km/일",
     "국토교통부 「2023 자동차 주행거리 통계」 비사업용 승용 평균 12,045 km/년"),
    ("평균 연비",                    12.5,    "km/L",
     "에너지관리공단 자동차 평균 연비 (휘발유 승용차)"),
    ("휘발유 단가",                  1_700,   "원/L",
     "오피넷(opinet.co.kr) 전국 평균가 (2026.05 기준)"),
    ("휘발유 CO₂ 배출계수",          2.31,    "kgCO₂/L",
     "환경부 국가 온실가스 인벤토리 — 휘발유 연소 계수"),
    ("시내버스 CO₂ 배출계수",        0.0265,  "kgCO₂/인·km",
     "환경부 교통수단별 온실가스 배출량 — 시내버스 인km당 평균"),
    ("연간 일수",                    365,     "일",
     "—"),
]

SCENARIOS = [0.05, 0.10, 0.15, 0.20]   # 5%, 10%, 15%, 20%

# 추출
CARS_TOTAL = 697_261
PASSENGER_RATIO = 0.80
CARS = round(CARS_TOTAL * PASSENGER_RATIO)             # 자가용 승용차
DAILY_KM_PER_CAR = 33.0
FUEL_EFF = 12.5
FUEL_PRICE = 1_700
GASOLINE_CO2 = 2.31
BUS_CO2 = 0.0265
DAYS = 365


# ──────────────────────────────────────────────────────────────────────
# 2. 시나리오 계산
# ──────────────────────────────────────────────────────────────────────
def compute_scenario(rate: float) -> dict:
    """전환율 → 절감 지표."""
    total_daily_km = CARS * DAILY_KM_PER_CAR              # 창원 자가용 일평균 총 통행거리(km)
    saved_km_daily = total_daily_km * rate                # 자가용에서 줄어든 km
    saved_l_daily = saved_km_daily / FUEL_EFF             # 절감 휘발유(L)
    saved_won_daily = saved_l_daily * FUEL_PRICE          # 절감 유류비(원)

    co2_car_per_km = GASOLINE_CO2 / FUEL_EFF              # 자가용 kgCO₂/km
    co2_saved_daily_kg = saved_km_daily * (co2_car_per_km - BUS_CO2)

    return {
        "rate_pct":           rate * 100,
        "saved_km_year":      saved_km_daily * DAYS,
        "saved_l_year":       saved_l_daily * DAYS,
        "saved_won_year":     saved_won_daily * DAYS,
        "saved_eok_year":     saved_won_daily * DAYS / 1e8,   # 억원
        "co2_saved_ton_year": co2_saved_daily_kg * DAYS / 1000,  # 톤
        "total_daily_km":     total_daily_km,
    }


# ──────────────────────────────────────────────────────────────────────
# 3. Excel 생성
# ──────────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="305496")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HIGHLIGHT_FILL = PatternFill("solid", fgColor="FFF2CC")
TITLE_FONT = Font(bold=True, size=14, color="1F3864")


def build_assumptions_sheet(ws):
    ws.title = "가정·출처"
    ws["A1"] = "차트 입력 가정 및 공공통계 출처"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    headers = ["항목", "값", "단위", "출처"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for i, (label, val, unit, src) in enumerate(ASSUMPTIONS, start=4):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=val)
        ws.cell(row=i, column=3, value=unit)
        ws.cell(row=i, column=4, value=src)

    widths = [28, 14, 12, 70]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def build_results_sheet(ws, scenarios: list[dict]):
    ws.title = "시나리오 결과"
    ws["A1"] = "대중교통 전환율별 연간 절감 효과"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:G1")

    headers = [
        "전환율 (%)",
        "자가용 절감 통행 (km/년)",
        "휘발유 절감 (L/년)",
        "유류비 절감 (원/년)",
        "유류비 절감 (억원/년)",
        "CO₂ 순감축 (톤/년)",
        "산정식 비고",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[3].height = 32

    notes = [
        "자가용 → 시내버스 전환 5%",
        "자가용 → 시내버스 전환 10%",
        "자가용 → 시내버스 전환 15%",
        "자가용 → 시내버스 전환 20%",
    ]
    for i, (s, note) in enumerate(zip(scenarios, notes), start=4):
        ws.cell(row=i, column=1, value=s["rate_pct"]).number_format = "0.0\"%\""
        ws.cell(row=i, column=2, value=s["saved_km_year"]).number_format = "#,##0"
        ws.cell(row=i, column=3, value=s["saved_l_year"]).number_format = "#,##0"
        ws.cell(row=i, column=4, value=s["saved_won_year"]).number_format = "#,##0"
        ws.cell(row=i, column=5, value=s["saved_eok_year"]).number_format = "#,##0.0"
        ws.cell(row=i, column=6, value=s["co2_saved_ton_year"]).number_format = "#,##0"
        ws.cell(row=i, column=7, value=note)

    # 핵심 컬럼 강조 (유류비·CO₂)
    for r in range(4, 4 + len(scenarios)):
        ws.cell(row=r, column=5).fill = HIGHLIGHT_FILL
        ws.cell(row=r, column=6).fill = HIGHLIGHT_FILL

    widths = [13, 24, 18, 22, 22, 20, 28]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws["A10"] = "산식"
    ws["A10"].font = Font(bold=True, color="1F3864")
    ws["A11"] = "자가용 절감 통행 = 자가용 등록대수 × 1대 일평균 주행 × 전환율 × 365"
    ws["A12"] = "휘발유 절감 = 자가용 절감 통행 ÷ 평균 연비(12.5 km/L)"
    ws["A13"] = "유류비 절감 = 휘발유 절감 × 휘발유 단가(1,700 원/L)"
    ws["A14"] = "CO₂ 순감축 = 자가용 절감 통행 × (휘발유 CO₂/연비 − 시내버스 CO₂)"


def build_chart_sheet(ws, scenarios: list[dict], src_ws_title: str):
    ws.title = "차트"
    ws["A1"] = "전환율 시나리오별 절감 효과"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    # 차트용 미니 데이터 테이블
    ws["A3"] = "전환율"
    ws["B3"] = "유류비 절감 (억원/년)"
    ws["C3"] = "CO₂ 순감축 (톤/년)"
    for cell in (ws["A3"], ws["B3"], ws["C3"]):
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for i, s in enumerate(scenarios, start=4):
        ws.cell(row=i, column=1, value=f"{s['rate_pct']:.0f}%")
        ws.cell(row=i, column=2, value=round(s["saved_eok_year"], 1))
        ws.cell(row=i, column=3, value=round(s["co2_saved_ton_year"], 0))

    # 유류비 차트
    chart_fuel = BarChart()
    chart_fuel.type = "col"
    chart_fuel.style = 11
    chart_fuel.title = "연간 유류비 절감 (억원)"
    chart_fuel.y_axis.title = "억원/년"
    chart_fuel.x_axis.title = "대중교통 전환율"
    data = Reference(ws, min_col=2, min_row=3, max_row=3 + len(scenarios), max_col=2)
    cats = Reference(ws, min_col=1, min_row=4, max_row=3 + len(scenarios))
    chart_fuel.add_data(data, titles_from_data=True)
    chart_fuel.set_categories(cats)
    chart_fuel.dLbls = DataLabelList(showVal=True)
    chart_fuel.height = 10
    chart_fuel.width = 18
    ws.add_chart(chart_fuel, "E3")

    # CO₂ 차트
    chart_co2 = BarChart()
    chart_co2.type = "col"
    chart_co2.style = 12
    chart_co2.title = "연간 CO₂ 순감축 (톤)"
    chart_co2.y_axis.title = "톤/년"
    chart_co2.x_axis.title = "대중교통 전환율"
    data2 = Reference(ws, min_col=3, min_row=3, max_row=3 + len(scenarios), max_col=3)
    chart_co2.add_data(data2, titles_from_data=True)
    chart_co2.set_categories(cats)
    chart_co2.dLbls = DataLabelList(showVal=True)
    chart_co2.height = 10
    chart_co2.width = 18
    ws.add_chart(chart_co2, "E23")

    widths = [12, 24, 22]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def build_excel(scenarios: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws_assume = wb.active
    build_assumptions_sheet(ws_assume)
    ws_result = wb.create_sheet()
    build_results_sheet(ws_result, scenarios)
    ws_chart = wb.create_sheet()
    build_chart_sheet(ws_chart, scenarios, ws_result.title)

    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)


# ──────────────────────────────────────────────────────────────────────
# 4. PNG (보고서 삽입용)
# ──────────────────────────────────────────────────────────────────────
def build_png(scenarios: list[dict]) -> None:
    # 한글 폰트 — 윈도우 맑은고딕
    plt.rcParams["font.family"] = "Malgun Gothic"
    plt.rcParams["axes.unicode_minus"] = False

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(13, 5))

    labels = [f"{s['rate_pct']:.0f}%" for s in scenarios]
    fuel_eok = [s["saved_eok_year"] for s in scenarios]
    co2_ton = [s["co2_saved_ton_year"] for s in scenarios]

    bars1 = ax1.bar(labels, fuel_eok,
                    color=["#A9D08E", "#70AD47", "#548235", "#375623"])
    ax1.set_title("연간 유류비 절감", fontsize=14, fontweight="bold", pad=12)
    ax1.set_xlabel("대중교통 전환율", fontsize=11)
    ax1.set_ylabel("억원 / 년", fontsize=11)
    ax1.yaxis.set_major_formatter(mtick.FuncFormatter(lambda x, _: f"{x:,.0f}"))
    ax1.grid(axis="y", alpha=0.3, linestyle="--")
    for b, v in zip(bars1, fuel_eok):
        ax1.text(b.get_x() + b.get_width() / 2, v, f"{v:,.0f}억",
                 ha="center", va="bottom", fontsize=10, fontweight="bold")

    bars2 = ax2.bar(labels, co2_ton,
                    color=["#9DC3E6", "#5B9BD5", "#2E75B6", "#1F4E79"])
    ax2.set_title("연간 CO₂ 순감축", fontsize=14, fontweight="bold", pad=12)
    ax2.set_xlabel("대중교통 전환율", fontsize=11)
    ax2.set_ylabel("톤 / 년", fontsize=11)
    ax2.yaxis.set_major_formatter(mtick.FuncFormatter(lambda x, _: f"{x:,.0f}"))
    ax2.grid(axis="y", alpha=0.3, linestyle="--")
    for b, v in zip(bars2, co2_ton):
        ax2.text(b.get_x() + b.get_width() / 2, v, f"{v:,.0f}톤",
                 ha="center", va="bottom", fontsize=10, fontweight="bold")

    fig.suptitle("창원시 대중교통 전환 시나리오별 절감 효과",
                 fontsize=16, fontweight="bold")
    fig.text(0.5, 0.01,
             "출처: 국토교통부 자동차등록현황보고·자동차주행거리통계 / 환경부 온실가스 인벤토리 / 오피넷 전국평균가",
             ha="center", fontsize=8, color="#666")
    plt.tight_layout(rect=[0, 0.03, 1, 0.95])
    plt.savefig(OUT_PNG, dpi=150, bbox_inches="tight")
    plt.close()


# ──────────────────────────────────────────────────────────────────────
# 5. main
# ──────────────────────────────────────────────────────────────────────
def main() -> int:
    print("[1] 시나리오 계산 (5/10/15/20%)")
    scenarios = [compute_scenario(r) for r in SCENARIOS]
    print(f"    창원 자가용 승용 {CARS:,} 대 × {DAILY_KM_PER_CAR} km/일 "
          f"= {scenarios[0]['total_daily_km']:,.0f} km/일 (총 자가용 통행)")

    print("\n[2] 결과 요약")
    print(f"    {'전환율':>6}  {'유류비 절감':>16}  {'CO₂ 순감축':>14}")
    print("    " + "-" * 44)
    for s in scenarios:
        print(f"    {s['rate_pct']:>5.0f}%  "
              f"{s['saved_eok_year']:>12,.1f} 억원  "
              f"{s['co2_saved_ton_year']:>10,.0f} 톤")

    print(f"\n[3] Excel 생성: {OUT_XLSX.relative_to(PROJECT_ROOT)}")
    build_excel(scenarios)
    print(f"    크기: {OUT_XLSX.stat().st_size / 1024:.1f} KB")

    print(f"\n[4] PNG 생성: {OUT_PNG.relative_to(PROJECT_ROOT)}")
    build_png(scenarios)
    print(f"    크기: {OUT_PNG.stat().st_size / 1024:.1f} KB")

    return 0


if __name__ == "__main__":
    sys.exit(main())
